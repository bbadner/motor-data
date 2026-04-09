import os
import re
import statistics
from datetime import datetime, timedelta
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from fpdf import FPDF
import traceback
import sys
import tkinter as tk
from tkinter import messagebox
import subprocess
from collections import defaultdict

# ======================================================
# CONFIGURATION
# ======================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_DIR = os.path.join(BASE_DIR, "Motor Reports")
LOGO_FILE = os.path.join(BASE_DIR, "rexair_logo.png")
LOG_FILE = os.path.join(BASE_DIR, "error_log.txt")
os.makedirs(REPORT_DIR, exist_ok=True)

def log_error(message):
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {message}\n")

# ======================================================
# HELPER FUNCTIONS
# ======================================================

def extract_date_from_filename(filename):
    """Extract date from pattern like (C6521250923-0001~6582)."""
    match = re.search(r"C6521(\d{2})(\d{2})(\d{2})", filename)
    if match:
        yy, mm, dd = match.groups()
        try:
            year = 2000 + int(yy)
            return datetime(year, int(mm), int(dd))
        except ValueError:
            log_error(f"Invalid date in filename: {filename}")
            return None
    return None

def read_input_power_data(filepath):
    """Read 'Input Power(W)' data under 'High Speed(Open)' section."""
    try:
        wb = load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
            text_rows = [" ".join(str(c or "").lower() for c in row) for row in rows]

            for i, line in enumerate(text_rows):
                if "high speed" in line and "open" in line:
                    for j in range(i, min(i + 30, len(rows))):
                        for k, cell in enumerate(rows[j]):
                            if cell and isinstance(cell, str) and "input power" in cell.lower():
                                data = []
                                for r in range(j + 1, len(rows)):
                                    val = rows[r][k]
                                    if isinstance(val, (int, float)):
                                        data.append(float(val))
                                if data:
                                    return data
        log_error(f"No 'Input Power(W)' data found in {os.path.basename(filepath)}")
        return []
    except Exception as e:
        log_error(f"Error reading {filepath}: {e}")
        return []

def compute_stats(values):
    """Return mean and standard deviation."""
    if not values:
        return None, None
    try:
        return statistics.mean(values), statistics.stdev(values)
    except Exception as e:
        log_error(f"Stats computation failed: {e}")
        return None, None

# ======================================================
# MAIN EXECUTION
# ======================================================

try:
    records = []
    latest_file = None
    latest_date = None

    for fname in os.listdir(BASE_DIR):
        if not fname.lower().endswith(".xlsx"):
            continue
        date = extract_date_from_filename(fname)
        if not date:
            continue
        fpath = os.path.join(BASE_DIR, fname)
        data = read_input_power_data(fpath)
        mean, std = compute_stats(data)
        if mean is not None:
            records.append((date, mean, std))
            if not latest_date or date > latest_date:
                latest_date = date
                latest_file = fpath

    if not records:
        raise SystemExit("❌ No valid motor data files found in folder.")

    records.sort(key=lambda x: x[0])
    start_date = latest_date - timedelta(days=365)
    records = [r for r in records if r[0] >= start_date]

    if not records:
        raise SystemExit("⚠️ No data found in the last 12 months.")

    # ======================================================
    # MONTHLY BOXPLOT + MEDIAN TREND + MEAN TREND + STD DEV
    # ======================================================

    monthly_data = defaultdict(list)
    for fname in os.listdir(BASE_DIR):
        if not fname.lower().endswith(".xlsx"):
            continue
        file_date = extract_date_from_filename(fname)
        if not file_date:
            continue
        if file_date < start_date or file_date > latest_date:
            continue
        data = read_input_power_data(os.path.join(BASE_DIR, fname))
        if data:
            key = file_date.strftime("%Y-%m")  # Year-month grouping
            monthly_data[key].extend(data)

    sorted_months = sorted(monthly_data.keys())
    month_labels = [datetime.strptime(m, "%Y-%m") for m in sorted_months]
    month_values = [monthly_data[m] for m in sorted_months]
    month_medians = [statistics.median(v) if len(v) > 0 else 0 for v in month_values]
    month_means = [statistics.mean(v) if len(v) > 0 else 0 for v in month_values]
    month_std = [statistics.stdev(v) if len(v) > 1 else 0 for v in month_values]

    fig, axes = plt.subplots(2, 1, figsize=(10, 8), sharex=True)
    fig.suptitle("Rexair Motor Performance Report - 12 Month Trend", fontsize=14, fontweight="bold")

    # --- Top Chart: Box & Whisker + Median & Mean ---
    positions = list(range(1, len(sorted_months) + 1))
    box = axes[0].boxplot(month_values, positions=positions, showfliers=True, patch_artist=True, widths=0.6)
    for patch in box['boxes']:
        patch.set(facecolor='lightblue', alpha=0.7)

    # Primary Y (Median)
    axes[0].plot(positions, month_medians, marker="o", color="darkblue", label="Median")

    # Secondary Y (Mean)
    ax2 = axes[0].twinx()
    ax2.plot(positions, month_means, marker="s", linestyle="--", color="orange", label="Mean (Monthly)")
    ax2.set_ylabel("Monthly Mean (W)", color="orange")

    axes[0].set_ylabel("Input Power (W)")
    axes[0].set_title("Monthly Input Power Distribution (High Speed - Open)")
    axes[0].grid(True, linestyle="--", alpha=0.6)
    axes[0].legend(loc="upper left")
    ax2.legend(loc="upper right")

    axes[0].set_xticks(positions)
    axes[0].set_xticklabels([d.strftime("%b %Y") for d in month_labels], rotation=45, ha="right")

    # --- Fixed Y-axis for clear separation ---
    axes[0].set_ylim(700, 1100)
    ax2.set_ylim(700, 1100)

    # --- Bottom Chart: Monthly Std Dev ---
    axes[1].plot(positions, month_std, marker="o", color="tab:orange", label="Standard Deviation")
    axes[1].set_ylabel("Std Dev (W)")
    axes[1].set_xlabel("Month")
    axes[1].legend()
    axes[1].grid(True, linestyle="--", alpha=0.6)
    axes[1].set_xticks(positions)
    axes[1].set_xticklabels([d.strftime("%b %Y") for d in month_labels], rotation=45, ha="right")

    plt.tight_layout(rect=[0, 0, 1, 0.96])
    chart_path = os.path.join(REPORT_DIR, "motor_trend_charts.png")
    plt.savefig(chart_path, dpi=200, bbox_inches="tight")
    plt.close()

    # ======================================================
    # SPC HISTOGRAM (±3σ VISUAL)
    # ======================================================

    latest_data = read_input_power_data(latest_file)
    hist_path = None
    if latest_data:
        mean = statistics.mean(latest_data)
        stdev = statistics.stdev(latest_data)
        lsl = mean - 3 * stdev
        usl = mean + 3 * stdev

        plt.figure(figsize=(8, 5))
        plt.hist(latest_data, bins=25, color="lightblue", edgecolor="black", alpha=0.7)
        plt.axvline(mean, color="blue", linestyle="-", linewidth=2, label=f"Mean = {mean:.2f}")
        plt.axvline(lsl, color="red", linestyle="--", linewidth=1.5, label=f"LSL (μ-3σ) = {lsl:.2f}")
        plt.axvline(usl, color="red", linestyle="--", linewidth=1.5, label=f"USL (μ+3σ) = {usl:.2f}")
        plt.title(f"Input Power (W) Distribution with Cp/Cpk Visuals\n{os.path.basename(latest_file)}")
        plt.xlabel("Input Power (W)")
        plt.ylabel("Frequency")
        plt.legend()
        plt.grid(True, linestyle="--", alpha=0.6)
        plt.tight_layout()

        hist_path = os.path.join(REPORT_DIR, "motor_histogram_spc.png")
        plt.savefig(hist_path, dpi=200)
        plt.close()

    # ======================================================
    # PDF REPORT
    # ======================================================

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    date_tag = datetime.now().strftime("%Y-%m-%d")
    report_pdf = os.path.join(REPORT_DIR, f"Rexair_Motor_Performance_Report_{date_tag}.pdf")

    class PDF(FPDF):
        def header(self):
            if os.path.exists(LOGO_FILE):
                self.image(LOGO_FILE, x=65, y=8, w=80)
                self.ln(25)
            self.set_font("Arial", "B", 14)
            self.cell(0, 10, "Rexair Motor Performance Report - 12 Month Trend", 0, 1, "C")
            self.set_font("Arial", "", 11)
            self.cell(0, 10,
                f"Period: {start_date.strftime('%b %Y')} - {latest_date.strftime('%b %Y')}", 0, 1, "C")
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "I", 8)
            self.cell(0, 10, f"Generated {timestamp}", 0, 0, "C")

    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Summary Statistics", 0, 1)
    pdf.set_font("Arial", "", 11)
    pdf.cell(40, 8, "Date", 1, 0, "C")
    pdf.cell(50, 8, "Mean Watts", 1, 0, "C")
    pdf.cell(50, 8, "Std Dev (W)", 1, 1, "C")

    for date, mean, std in records:
        pdf.cell(40, 8, date.strftime("%b %d, %Y"), 1, 0, "C")
        pdf.cell(50, 8, f"{mean:.2f}", 1, 0, "C")
        pdf.cell(50, 8, f"{std:.2f}", 1, 1, "C")

    pdf.ln(10)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Trend Charts", 0, 1)
    pdf.image(chart_path, w=180)

    if hist_path and os.path.exists(hist_path):
        pdf.add_page()
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, "Latest Input Power (W) SPC Histogram (+/-3 sigma)", 0, 1, "C")
        pdf.image(hist_path, w=180)

    pdf_output_bytes = bytes(str(pdf.output(dest='S')), "latin-1", errors="replace")
    with open(report_pdf, "wb") as f:
        f.write(pdf_output_bytes)

    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Rexair Report Generator", f"✅ Report created successfully!\n\nSaved to:\n{report_pdf}")

    try:
        os.startfile(report_pdf)
    except Exception:
        subprocess.Popen(["open", report_pdf])

    sys.exit(0)

except Exception as e:
    error_details = traceback.format_exc()
    log_error(f"FATAL ERROR: {e}\n{error_details}")
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Rexair Report Generator", "❌ A fatal error occurred.\nPlease check error_log.txt for details.")
    sys.exit(1)
